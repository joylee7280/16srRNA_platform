import subprocess as sp
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime
def change_d(file):
    fin = open(file, "rt")
    line = fin.read()
    line = line.replace('hey', 'hello')
    line = line.replace('<link rel="stylesheet" href="./q2templateassets/css/bootstrap.min.css" />',
                        '<link rel="stylesheet" href="{% static \'q2templateassets/css/bootstrap.min.css\' %}"/>')
    line = line.replace('href="./q2templateassets/css/bootstrap.min.css"',
                        'href="{% static \'q2templateassets/css/bootstrap.min.css\' %}"')
    line = line.replace('href="./q2templateassets/css/normalize.css"/',
                        'href="{% static \'q2templateassets/css/normalize.css\' %}"')
    line = line.replace('href="./q2templateassets/css/base-template.css"',
                        'href="{% static \'q2templateassets/css/base-template.css\' %}"')
    line = line.replace('src="./q2templateassets/js/jquery-3.2.0.min.js"',
                        'src="{% static \'q2templateassets/js/jquery-3.2.0.min.js\' %}"')
    line = line.replace('src="./q2templateassets/js/bootstrap.min.js"',
                        'src="{% static \'q2templateassets/js/bootstrap.min.js\' %}"')
    line = line.replace('<img src="./q2templateassets/img/qiime2-rect-200.png"',
                        '<img src="{% static \'q2templateassets/img/qiime2-rect-200.png\' %}"')
    fin.close()

    fin = open(file, "wt")
    fin.write(line)
    fin.close()

    fin = open(file, "rt")
    rows = fin.readlines()
    rows.insert(4, '{% load static %}\n')
    fin.close()

    fin = open(file, "wt")
    for row in rows:
        fin.write(row)
    fin.close()
def change_a(file,user,project):
    stri = "$('<a href=\"data.tsv\" target=\"_blank\" rel=\"noopener noreferrer\" class=\"btn btn-default\">Export as TSV</a>')"
    stri_fix = "$('<a href=\"/ancom/ancom/download_data/\" target=\"_blank\" rel=\"noopener noreferrer\" class=\"btn btn-default\">Export as TSV</a>')"
    fin = open(file, "rt")
    line = fin.read()
    line = line.replace('<script src="js/vega.min.js"></script>',
                        '<script src="{% static \'js/vega.min.js\' %}"></script>')
    line = line.replace('<script src="js/vega-embed.min.js"></script>',
                        '<script src="{% static \'js/vega-embed.min.js\' %}"></script>')
    line = line.replace('<link rel="stylesheet" type="text/css" href="css/spinkit.css">',
                        '<link rel="stylesheet" type="text/css" href="{% static \'css/spinkit.css\' %}">')
    line = line.replace('<a href="ancom.tsv" target="_blank" rel="noopener noreferrer" class="btn btn-default">',
                        '<a href="/ancom/ancom/download_ancom/" target="_blank" rel="noopener noreferrer" class="btn btn-default">')
    line = line.replace('<a href="percent-abundances.tsv" target="_blank" rel="noopener noreferrer" class="btn btn-default">',
                        '<a href="/ancom/ancom/download_abun/" target="_blank" rel="noopener noreferrer" class="btn btn-default">')
    line = line.replace(stri,stri_fix)
    fin.close()

    fin = open(file, "wt")
    fin.write(line)
    fin.close()
def change_d_barplot(file,user,project):
    fin = open(file, "rt")
    line = fin.read()
    line = line.replace('<li><a href="./bodysiteleft%2520palm-ancombc-barplot.html">bodysiteleft palm</a></li>',
                        '<li><a href="/work1791/b09901010/project/mainsite/templates/user_output/'+user+'/'+project+'/ancombc/differentials_barplot_unzip/bodysiteleft%2520palm-ancombc-barplot.html">bodysiteleft palm</a></li>')
    line = line.replace('<li><a href="./bodysiteright%2520palm-ancombc-barplot.html">bodysiteright palm</a></li>',
                        '<li><a href="/work1791/b09901010/project/mainsite/templates/user_output/'+user+'/'+project+'/ancombc/differentials_barplot_unzip/bodysiteright%2520palm-ancombc-barplot.html">bodysiteright palm</a></li>')
    line = line.replace('<li><a href="./bodysitetongue-ancombc-barplot.html">bodysitetongue</a></li>',
                        '<li><a href="/work1791/b09901010/project/mainsite/templates/user_output/'+user+'/'+project+'/ancombc/differentials_barplot_unzip/bodysitetongue-ancombc-barplot.html">bodysitetongue</a></li>')
    fin.close()

    fin = open(file, "wt")
    fin.write(line)
    fin.close()

#ancom
def vis_ancom(user,project,metadata,table,taxonomy,clas,para,metadata_col):
    sp.run(["qiime feature-table filter-samples \
    --i-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/table/"+table+" \
    --m-metadata-file /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/metadata/"+metadata+" \
    --p-where \"["+clas+"]=\'"+para+"\'\" \
    --o-filtered-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/gut-table.qza"], shell=True,capture_output=True, text=True)

    sp.run(["qiime taxa collapse \
    --i-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/gut-table.qza \
    --i-taxonomy /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/taxonomy/"+taxonomy+" \
    --p-level 6 \
    --o-collapsed-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/collapse.table.qza"], shell=True,capture_output=True, text=True)
    sp.run(["mkdir /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancom"], shell=True,capture_output=True, text=True)
    sp.run(["mkdir /work1791/b09901010/project/mainsite/templates/user_output/"+user+"/"+project+"/ancom"], shell=True,capture_output=True, text=True)
    ################ log ################
    sp.run(["touch /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancom/ancom.txt"], shell=True,
                                        capture_output=True, text=True)
    path = "/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancom/ancom.txt"
    with open(path, 'a') as f:
        f.write("Folder data\n")
        f.write("# user:"+user+"\n")
        f.write("# project:"+project+"\n")
        f.write("# analysis:ancom\n")
    #####################################
    #####################################
    with open(path, 'a') as f:
        f.write("# class:"+str(clas)+"\n")
        f.write("# parameter:"+str(para)+"\n")
    #####################################
    sp.run(["qiime composition add-pseudocount \
    --i-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/collapse.table.qza \
    --o-composition-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancom/ancom.qza"], shell=True,capture_output=True, text=True)
    sp.run(["qiime composition ancom \
    --i-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancom/ancom.qza \
    --m-metadata-file /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/metadata/"+metadata+" \
    --m-metadata-column "+metadata_col+" \
    --o-visualization /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancom/ancom.qzv"], shell=True,capture_output=True, text=True)
    #####################################
    with open(path, 'a') as f:
        f.write("# metadata_coloumn:"+str(metadata_col)+"\n")
        f.write("\n")
    #####################################
    #####################################
    with open(path, 'a') as f:
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" \"ancom.qzv\"|builded\n")
    print(datetime.datetime.now())
    #####################################
    sp.run(["qiime tools export \
    --input-path /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancom/ancom.qzv \
    --output-path /work1791/b09901010/project/mainsite/templates/user_output/"+user+"/"+project+"/ancom/ancom_unzip"], shell=True,capture_output=True, text=True)
    #####################################
    with open(path, 'a') as f:
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" \"ancom_unzip\"|exported\n")
    print(datetime.datetime.now())
    #####################################
    change_d("/work1791/b09901010/project/mainsite/templates/user_output/"+user+"/"+project+"/ancom/ancom_unzip/index.html")
    change_a("/work1791/b09901010/project/mainsite/templates/user_output/"+user+"/"+project+"/ancom/ancom_unzip/index.html",user,project)
    ################ log ################
    with open(path, 'a') as f:
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" !!! finish !!!\n")
    #####################################
    ################ log ################
    path = "/work1791/b09901010/project/user_folder/"+user+"/"+project+"/project.txt"
    with open(path, 'a') as f:
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" !!! ancom|finish !!!\n")
    #####################################
#ancombc
def vis_ancombc(user,project,p_formula):
    sp.run(["qiime taxa collapse \
    --i-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/table/table.qza \
    --i-taxonomy /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/taxonomy/taxonomy.qza \
    --p-level 6 \
    --o-collapsed-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/collapse.table.qza"], shell=True,capture_output=True, text=True)
    sp.run(["mkdir /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancombc"], shell=True,capture_output=True, text=True)
    ################ log ################
    sp.run(["touch /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancombc/ancombc.txt"], shell=True,
                                        capture_output=True, text=True)
    path = "/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancombc/ancombc.txt"
    with open(path, 'a') as f:
        f.write("Folder data\n")
        f.write("# user:"+user+"\n")
        f.write("# project:"+project+"\n")
        f.write("# analysis:ancombc\n")
    #####################################
    sp.run(["mkdir /work1791/b09901010/project/mainsite/templates/user_output/"+user+"/"+project+"/ancombc"], shell=True,capture_output=True, text=True)
    sp.run(["qiime composition ancombc \
        --i-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/collapse.table.qza \
        --m-metadata-file /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/metadata/sample-metadata.tsv \
        --p-formula "+p_formula+" \
        --o-differentials /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancombc/dataloaf.qza"], shell=True,capture_output=True, text=True)
    ################ log ################
    with open(path, 'a') as f:
        f.write("# class:"+str(p_formula)+"\n")
        f.write("\n")
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" \"dataloaf.qza\"|builded\n")
    #####################################
    sp.run(["qiime composition tabulate \
    --i-data /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output//ancombc/dataloaf.qza \
    --o-visualization /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancombc/differentials.qzv"], shell=True,capture_output=True, text=True)
    sp.run(["qiime composition da-barplot \
    --i-data /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancombc/dataloaf.qza \
    --o-visualization /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancombc/differentials_barplot.qzv"], shell=True,capture_output=True, text=True)
    ################ log ################
    with open(path, 'a') as f:
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" \"differentials_barplot.qzv\"|builded\n")
    #####################################
    sp.run(["qiime tools export \
    --input-path /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancombc/differentials.qzv \
    --output-path /work1791/b09901010/project/mainsite/templates/user_output/"+user+"/"+project+"/ancombc/differentials_unzip"], shell=True,capture_output=True, text=True)
    sp.run(["qiime tools export \
    --input-path /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/ancombc/differentials_barplot.qzv \
    --output-path /work1791/b09901010/project/mainsite/templates/user_output/"+user+"/"+project+"/ancombc/differentials_barplot_unzip"], shell=True,capture_output=True, text=True)

    change_d("/work1791/b09901010/project/mainsite/templates/user_output/"+user+"/"+project+"/ancombc/differentials_unzip/index.html")
    change_d_barplot("/work1791/b09901010/project/mainsite/templates/user_output/"+user+"/"+project+"/ancombc/differentials_unzip/index.html",user,project)
    ################ log ################
    with open(path, 'a') as f:
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" !!! finish !!!\n")
    #####################################
    ################ log ################
    path = "/work1791/b09901010/project/user_folder/"+user+"/"+project+"/"+str(project)+".txt"
    with open(path, 'a') as f:
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" !!! ancombc|finish !!!\n")
    #####################################
#lefse_convert
def vis_lefse_convert(user,project):
    sp.run(["qiime taxa collapse \
    --i-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/table/table.qza \
    --i-taxonomy /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/taxonomy/taxonomy.qza \
    --p-level 6 \
    --o-collapsed-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/collapse.table.qza"], shell=True,capture_output=True, text=True)

    sp.run(["mkdir /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe"], shell=True,capture_output=True, text=True)
    sp.run(["mkdir /work1791/b09901010/project/mainsite/templates/user_output/"+user+"/"+project+"/LEfSe"], shell=True,capture_output=True, text=True)
    sp.run(["mkdir /work1791/b09901010/project/mainsite/static/user_output/"], shell=True,capture_output=True, text=True)
    sp.run(["mkdir /work1791/b09901010/project/mainsite/static/user_output/"+user], shell=True,capture_output=True, text=True)
    sp.run(["mkdir /work1791/b09901010/project/mainsite/static/user_output/"+user+"/"+project], shell=True,capture_output=True, text=True)
    sp.run(["mkdir /work1791/b09901010/project/mainsite/static/user_output/"+user+"/"+project+"/LEfSe"], shell=True,capture_output=True, text=True)
    sp.run(["qiime feature-table relative-frequency \
        --i-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/collapse.table.qza \
        --o-relative-frequency-table /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/collapse.frequency.table.qza"], shell=True,capture_output=True, text=True)
    sp.run(["qiime tools export \
        --input-path /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/collapse.frequency.table.qza \
        --output-path /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/"], shell=True,capture_output=True, text=True)
    sp.run(["biom convert \
    -i /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/feature-table.biom \
    -o /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/collapse.frequency.table.tsv \
    --header-key “taxonomy” \
    --to-tsv"], shell=True,capture_output=True, text=True)
    ################ log ################
    sp.run(["touch /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/LEfSe.txt"], shell=True,
                                        capture_output=True, text=True)
    path = "/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/LEfSe.txt"
    with open(path, 'a') as f:
        f.write("Folder data\n")
        f.write("# user:"+user+"\n")
        f.write("# project:"+project+"\n")
        f.write("# analysis:LEfSe\n")
        f.write("\n")
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" original collapse table builed \n")
    #####################################
    sp.run(["sed -i 's/;/\|/g' /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/collapse.frequency.table.tsv"], shell=True,capture_output=True, text=True)
    sp.run(["sed -i '1d' /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/collapse.frequency.table.tsv"], shell=True,capture_output=True, text=True)
    sp.run(["sed -i 's/\“taxonomy\”//g' /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/collapse.frequency.table.tsv"], shell=True,capture_output=True, text=True)
    tsv_table = pd.read_csv("/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/collapse.frequency.table.tsv",
                            sep='\t')
    tsv_table.to_excel("/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/collapse.frequency.table.xlsx", index=False)
    tsv_meta = pd.read_csv("/work1791/b09901010/project/user_folder/"+user+"/"+project+"/input/metadata/sample-metadata.tsv", sep='\t')
    tsv_meta.to_excel("/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/sample.xlsx", index=False)
    exc_table = openpyxl.load_workbook("/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/collapse.frequency.table.xlsx")
    exc_table.save("/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/collapse.frequency.table.xlsx")
    exc_meta = openpyxl.load_workbook("/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/sample.xlsx")
    exc_meta.save("/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/sample.xlsx")
    ws1 = exc_table.active
    ws2 = exc_meta.active
    ws1.insert_rows(0)
    ws1['A1'] = "class"
    for i in range(2, ws1.max_column+1):
        for j in range(3, ws2.max_row+1):
            char = get_column_letter(i)
            if (ws1[char+"2"].value == ws2['A'+str(j)].value):
                ws1[char+"1"].value = ws2['C'+str(j)].value
                ws1[char+"1"].value = ws1[char+"1"].value.replace(" ", "_")
                break
    exc_table.save("/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/converted.xlsx")
    df = pd.read_excel("/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/converted.xlsx", sheet_name='Sheet1',
                    header=None)
    df.to_csv("/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/converted.tsv", sep="\t", header=False, index=False)
    #####################################
    with open(path, 'a') as f:
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" adjust collapse table \n")
    #####################################
#lefse
def vis_lefse(user,project):
    path = "/work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/LEfSe.txt"
    sp.run(["conda run -n python2 lefse_format_input.py \
        /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/converted.tsv \
        /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/converted.tsv.lefse.in \
        -c 1 \
        -u 2 \
        -m f \
        -o 100000"], shell=True,capture_output=True, text=True)
    sp.run(["conda run -n python2 lefse_run.py \
        /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/converted.tsv.lefse.in \
        /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/converted.tsv.lefse.res"], shell=True,capture_output=True, text=True)
    sp.run(["conda run -n python2 lefse_plot_res.py \
        /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/converted.tsv.lefse.res \
        /work1791/b09901010/project/mainsite/templates/user_output/"+user+"/"+project+"/LEfSe/lefse_final_lda.pdf \
        --format pdf \
        --autoscale 0"], shell=True,capture_output=True, text=True)
    #####################################
    with open(path, 'a') as f:
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" LEfSe lda plot|builded \n")
    #####################################
    sp.run(["conda run -n python2 lefse_plot_res.py \
        /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/converted.tsv.lefse.res \
        /work1791/b09901010/project/mainsite/static/user_output/"+user+"/"+project+"/LEfSe/lefse_final_lda.png \
        --format png \
        --dpi 330 \
        --autoscale 0"], shell=True,capture_output=True, text=True)
    sp.run(["conda run -n python2 lefse_plot_cladogram.py \
        /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/converted.tsv.lefse.res \
        /work1791/b09901010/project/mainsite/templates/user_output/"+user+"/"+project+"/LEfSe/lefse_total_clado.pdf \
        --format pdf"], shell=True,capture_output=True, text=True)
    #####################################
    with open(path, 'a') as f:
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" LEfSe cladogram|builded \n")
    #####################################
    sp.run(["conda run -n python2 lefse_plot_cladogram.py \
        /work1791/b09901010/project/user_folder/"+user+"/"+project+"/output/LEfSe/converted.tsv.lefse.res \
        /work1791/b09901010/project/mainsite/static/user_output/"+user+"/"+project+"/LEfSe/lefse_total_clado.png \
        --format png \
        --dpi 330"], shell=True,capture_output=True, text=True)
    ################ log ################
    with open(path, 'a') as f:
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" !!! finish !!!\n")
    #####################################
    ################ log ################
    path = "/work1791/b09901010/project/user_folder/"+user+"/"+project+"/"+str(project)+".txt"
    with open(path, 'a') as f:
        f.write("["+str(datetime.datetime.now())+"]")
        f.write(" !!! LEfSe|finish !!!\n")
    #####################################
#correct
def vis_lefse_output(s_userID,ur_proj):
    file = "/work1791/b09901010/project/mainsite/templates/user_output/"+s_userID+"/"+ur_proj+"/LEfSe/lefse_output.html"
    lda = "{% static 'user_output/"+s_userID+"/"+ur_proj+"/LEfSe/lefse_final_lda.png'%}"
    clado = "{% static 'user_output/"+s_userID+"/"+ur_proj+"/LEfSe/lefse_total_clado.png'%}"
    fin = open(file, "rt")
    line = fin.read()
    line = line.replace('lda_insert',
                        lda)
    line = line.replace('clado_insert',
                        clado)
    fin.close()

    fin = open(file, "wt")
    fin.write(line)
    fin.close()
